"""Read the XLSX workbook inside a Stage 5 download zip."""

from __future__ import annotations

import argparse
import json
import zipfile
from io import BytesIO
from pathlib import Path
from typing import cast

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--zip-path", required=True)
    parser.add_argument("--sheet-name", required=True)
    args = parser.parse_args()

    with zipfile.ZipFile(Path(args.zip_path), "r") as archive:
        workbook_name = next(name for name in archive.namelist() if name.endswith(".xlsx"))
        workbook_bytes = BytesIO(archive.read(workbook_name))
    workbook = load_workbook(workbook_bytes, data_only=True)
    sheet = cast(Worksheet, workbook[args.sheet_name])
    rows = [
        [str(value) if value is not None else "" for value in row]
        for row in sheet.iter_rows(values_only=True)
    ]
    print(json.dumps(rows))


if __name__ == "__main__":
    main()
