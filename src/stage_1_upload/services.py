"""Analyze tabular structure and infer column types for upload preview.

Produces both the small ``ColumnPreview`` list (legacy 5-row sample for the
upload screen) and a ``ColumnProfile`` per column (full distinct-value tally
consumed by the Stage 2 takeover left pane).
"""

from __future__ import annotations

from collections.abc import Iterable
from datetime import datetime
from pathlib import Path
from typing import cast

from netrias_client import TabularColumn, read_tabular
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.domain.column_profile import ColumnProfile, build_column_profile
from src.domain.manifest import completeness_bucket

from .schemas import ColumnPreview, SheetPreview

DEFAULT_SHEET_PREVIEW_ROWS = 5
DEFAULT_SHEET_PREVIEW_COLUMNS = 6

# Workbook previews have one header row followed by the data rows shown on the
# upload page. Keep these names close to the slicing logic so the row math reads
# as UI behavior instead of OpenPyXL indexing trivia.
_HEADER_ROW_COUNT = 1
_HEADER_ROW_INDEX = 0
_FIRST_DATA_ROW_INDEX = _HEADER_ROW_INDEX + _HEADER_ROW_COUNT
_OPENPYXL_EMPTY_SHEET_SIZE = 1
_OPENPYXL_FIRST_CELL = "A1"


def analyze_columns(
    csv_path: Path,
    max_preview_rows: int = 5,
    sheet_name: str | None = None,
) -> tuple[int, list[ColumnPreview], dict[str, ColumnProfile]]:
    if not csv_path.exists():
        raise FileNotFoundError(csv_path)

    dataset = read_tabular(csv_path, sheet_name=sheet_name)
    profiles = {
        column.key: build_column_profile(
            column.key,
            (row[column.index] if column.index < len(row) else "" for row in dataset.rows),
        )
        for column in dataset.columns
    }
    sample_rows = dataset.rows[:max_preview_rows]
    columns = [
        _analyze_single_column(
            column,
            sample_rows,
            profiles[column.key],
        )
        for column in dataset.columns
    ]
    total_rows = len(dataset.rows)
    return total_rows, columns, profiles


def read_workbook_sheet_previews(
    path: Path,
    sheet_names: list[str],
    *,
    max_rows: int = DEFAULT_SHEET_PREVIEW_ROWS,
    max_cols: int = DEFAULT_SHEET_PREVIEW_COLUMNS,
) -> dict[str, SheetPreview]:
    """Return small exact-value worksheet previews for the upload response."""
    if path.suffix.lower() != ".xlsx" or not sheet_names:
        return {}
    if not path.exists():
        raise FileNotFoundError(path)

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        previews: dict[str, SheetPreview] = {}
        for sheet_name in sheet_names:
            worksheet = cast(Worksheet, workbook[sheet_name])
            previews[sheet_name] = _read_single_sheet_preview(worksheet, max_rows=max_rows, max_cols=max_cols)
        return previews
    finally:
        workbook.close()


def _read_single_sheet_preview(worksheet: Worksheet, *, max_rows: int, max_cols: int) -> SheetPreview:
    """Read only the worksheet rectangle the upload page can render.

    The worksheet dimensions already tell us whether rows or columns are hidden
    by the preview limit, so the data read stays limited to visible cells.
    """
    if _worksheet_is_empty(worksheet):
        return SheetPreview()

    visible_columns = min(worksheet.max_column, max_cols)
    visible_row_count = _HEADER_ROW_COUNT + max_rows
    rows = list(worksheet.iter_rows(max_row=visible_row_count, max_col=visible_columns, values_only=True))
    header_values = list(rows[_HEADER_ROW_INDEX] if rows else ())
    headers = [_cell_to_string(value) for value in header_values[:visible_columns]]
    preview_rows = [
        _shape_preview_row(row, width=len(headers))
        for row in rows[_FIRST_DATA_ROW_INDEX:visible_row_count]
    ]
    truncated_rows = worksheet.max_row > visible_row_count
    truncated_columns = worksheet.max_column > max_cols
    return SheetPreview(
        headers=headers,
        rows=preview_rows,
        truncated_rows=truncated_rows,
        truncated_columns=truncated_columns,
    )


def _worksheet_is_empty(worksheet: Worksheet) -> bool:
    return (
        worksheet.max_row == _OPENPYXL_EMPTY_SHEET_SIZE
        and worksheet.max_column == _OPENPYXL_EMPTY_SHEET_SIZE
        and worksheet[_OPENPYXL_FIRST_CELL].value is None
    )


def _shape_preview_row(row: tuple[object, ...], *, width: int) -> list[str]:
    values = [_cell_to_string(value) for value in row[:width]]
    return values + [""] * (width - len(values))


def _cell_to_string(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _analyze_single_column(
    column: TabularColumn,
    sample_rows: list[list[str]],
    profile: ColumnProfile,
) -> ColumnPreview:
    samples = [_normalize_sample(row[column.index] if column.index < len(row) else "") for row in sample_rows]
    non_empty_values = [value for value in samples if value]
    non_empty_count = len(non_empty_values)
    sample_size = max(len(samples), 1)

    return ColumnPreview(
        column_name=column.header,
        column_key=column.key,
        source_index=column.index,
        header=column.header,
        inferred_type=_infer_type(non_empty_values),
        sample_values=samples,
        has_non_empty_values=profile.total_distinct > 0,
        confidence_bucket=completeness_bucket(non_empty_count, sample_size),
        confidence_score=round(non_empty_count / sample_size, 2),
    )


def _normalize_sample(value: str | None) -> str:
    """Truncate to 80 chars to prevent oversized UI tooltips."""
    if value is None:
        return ""
    sanitized = value.strip()
    return sanitized[:80]


def _infer_type(values: Iterable[str]) -> str:
    cleaned = [value.replace(",", "") for value in values if value]
    if cleaned and _looks_numeric(cleaned):
        return "numeric"
    if cleaned and _looks_date(cleaned):
        return "date"
    if cleaned:
        return "text"
    return "unknown"


def _looks_numeric(values: list[str]) -> bool:
    try:
        for value in values:
            float(value)
        return True
    except ValueError:
        return False


def _looks_date(values: list[str]) -> bool:
    if not values:
        return False
    formats = ["%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d"]
    return all(_matches_any_date_format(candidate, formats) for candidate in values)


def _matches_any_date_format(value: str, formats: list[str]) -> bool:
    for fmt in formats:
        try:
            datetime.strptime(value, fmt)
            return True
        except ValueError:
            continue
    return False
