"""Stage 5 summary and download feature tests."""

from __future__ import annotations

import csv
import io
import json
import zipfile
from io import BytesIO
from typing import cast

import pytest
from httpx import AsyncClient
from netrias_client import TabularFormat, dataset_from_rows, write_tabular
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import src.app.dependencies as dependencies
from src.storage import UploadStorage, WorkflowFile
from tests.conftest import (
    TEST_TARGET_SCHEMA,
    TEST_TSV_CONTENT_TYPE,
    TEST_XLSX_CONTENT_TYPE,
    create_csv_content,
    create_harmonized_csv,
    create_manifest_for_file,
    create_xlsx_content,
    review_state_payload,
    store_test_completed_harmonization,
    upload_content,
)

pytestmark = pytest.mark.asyncio


def _read_downloaded_csv(response_bytes: bytes) -> list[dict[str, str]]:
    with zipfile.ZipFile(BytesIO(response_bytes), "r") as archive:
        csv_name = next(name for name in archive.namelist() if name.endswith(".csv"))
        csv_content = archive.read(csv_name).decode("utf-8")
    return list(csv.DictReader(io.StringIO(csv_content)))


def _read_downloaded_tabular(response_bytes: bytes, suffix: str, delimiter: str) -> list[list[str]]:
    with zipfile.ZipFile(BytesIO(response_bytes), "r") as archive:
        data_name = next(name for name in archive.namelist() if name.endswith(suffix))
        content = archive.read(data_name).decode("utf-8")
    return list(csv.reader(io.StringIO(content), delimiter=delimiter))


def _read_downloaded_xlsx(response_bytes: bytes, sheet_name: str) -> list[list[str]]:
    with zipfile.ZipFile(BytesIO(response_bytes), "r") as archive:
        workbook_name = next(name for name in archive.namelist() if name.endswith(".xlsx"))
        workbook_bytes = BytesIO(archive.read(workbook_name))
    workbook = load_workbook(workbook_bytes, data_only=True)
    sheet = cast(Worksheet, workbook[sheet_name])
    return [[str(value) if value is not None else "" for value in row] for row in sheet.iter_rows(values_only=True)]


async def test_stage5_download_matches_harmonized_when_no_overrides(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Download returns the harmonized dataset when no overrides exist."""

    # Given: an uploaded file with harmonized output and no overrides
    rows = [["col_a"], ["alpha"], ["beta"]]
    file_id = await upload_content(app_client, create_csv_content(rows), "download.csv")
    meta = temp_storage.load(file_id)
    assert meta is not None
    create_harmonized_csv(temp_storage, file_id, meta.saved_path, {1: {"col_a": "gamma"}})
    create_manifest_for_file(temp_storage, file_id, meta.saved_path, {1: {"col_a": "gamma"}})

    # When: the download endpoint is invoked
    response = await app_client.post("/stage-5/download", json={"file_id": file_id})

    # Then: the CSV reflects harmonized values
    assert response.status_code == 200
    output_rows = _read_downloaded_csv(response.content)
    assert output_rows[1]["col_a"] == "gamma"


async def test_stage5_download_preserves_harmonized_headers(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Download keeps headers produced by harmonization, including Stage 2 column renames."""

    # Given: harmonization wrote a renamed output column
    rows = [["diagnosis"], ["alpha"]]
    file_id = await upload_content(app_client, create_csv_content(rows), "renamed.csv")
    meta = temp_storage.load(file_id)
    assert meta is not None
    harmonized_path = temp_storage.harmonized_path_for(file_id, meta.saved_path)
    with harmonized_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerows([["Primary Diagnosis"], ["Lung Cancer"]])
    store_test_completed_harmonization(temp_storage, file_id, harmonized_path)

    # When: the download endpoint is invoked
    response = await app_client.post("/stage-5/download", json={"file_id": file_id})

    # Then: the downloaded CSV keeps the harmonized header and value
    assert response.status_code == 200
    output_rows = _read_downloaded_csv(response.content)
    assert list(output_rows[0]) == ["Primary Diagnosis"]
    assert output_rows[0]["Primary Diagnosis"] == "Lung Cancer"


async def test_stage5_download_succeeds_without_manifest(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Download succeeds even when a manifest is missing."""

    # Given: an uploaded file with a harmonized CSV but no manifest
    rows = [["col_a"], ["alpha"], ["beta"]]
    file_id = await upload_content(app_client, create_csv_content(rows), "no-manifest.csv")
    meta = temp_storage.load(file_id)
    assert meta is not None
    harmonized_path = create_harmonized_csv(temp_storage, file_id, meta.saved_path, {})
    store_test_completed_harmonization(temp_storage, file_id, harmonized_path)

    # When: the download endpoint is invoked
    response = await app_client.post("/stage-5/download", json={"file_id": file_id})

    # Then: the response contains only the CSV file
    assert response.status_code == 200
    with zipfile.ZipFile(BytesIO(response.content), "r") as archive:
        names = archive.namelist()
    assert any(name.endswith(".csv") for name in names)
    assert not any(name.endswith(".parquet") for name in names)


async def test_stage5_download_includes_manifest_json_when_available(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Download bundles the harmonization manifest as inspectable JSON."""

    # Given: an uploaded file with harmonized output and a stored manifest
    rows = [["col_a"], ["alpha"]]
    file_id = await upload_content(app_client, create_csv_content(rows), "with-manifest.csv")
    meta = temp_storage.load(file_id)
    assert meta is not None
    create_harmonized_csv(temp_storage, file_id, meta.saved_path, {0: {"col_a": "beta"}})
    create_manifest_for_file(temp_storage, file_id, meta.saved_path, {0: {"col_a": "beta"}})

    # When: the download endpoint is invoked
    response = await app_client.post("/stage-5/download", json={"file_id": file_id})

    # Then: the ZIP includes a JSON copy of the manifest rows
    assert response.status_code == 200
    with zipfile.ZipFile(BytesIO(response.content), "r") as archive:
        manifest_name = next(name for name in archive.namelist() if name.endswith("_manifest.json"))
        manifest_rows = json.loads(archive.read(manifest_name).decode("utf-8"))
    assert manifest_rows[0]["column_name"] == "col_a"
    assert manifest_rows[0]["to_harmonize"] == "alpha"
    assert manifest_rows[0]["top_harmonization"] == "beta"


async def test_stage5_download_includes_cde_mapping_artifact(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Download bundles the saved column-to-CDE mapping plan when available."""

    # Given: an uploaded file with a saved CDE mapping document
    rows = [["col_a"], ["alpha"]]
    file_id = await upload_content(app_client, create_csv_content(rows), "with-mapping.csv")
    meta = temp_storage.load(file_id)
    assert meta is not None
    harmonized_path = create_harmonized_csv(temp_storage, file_id, meta.saved_path, {})
    store_test_completed_harmonization(temp_storage, file_id, harmonized_path)
    dependencies.get_workflow_storage().write_json(
        dependencies.get_user_context(),
        file_id,
        WorkflowFile.CDE_MAPPING,
        {
            "file_id": file_id,
            "generated_at": "2026-05-13T00:00:00+00:00",
            "data_model_key": TEST_TARGET_SCHEMA,
            "external_version_number": "1",
            "mappings": [{"column_key": "col_0000", "source_column_name": "col_a"}],
        },
    )

    # When: the download endpoint is invoked
    response = await app_client.post("/stage-5/download", json={"file_id": file_id})

    # Then: the ZIP includes the mapping artifact alongside the data file
    assert response.status_code == 200
    with zipfile.ZipFile(BytesIO(response.content), "r") as archive:
        mapping_name = next(name for name in archive.namelist() if name.endswith("_cde_mapping.json"))
        mapping_document = json.loads(archive.read(mapping_name).decode("utf-8"))
    assert mapping_document["file_id"] == file_id
    assert mapping_document["mappings"][0]["column_key"] == "col_0000"


async def test_stage5_download_tsv_input_exports_tsv(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """A TSV upload downloads a TSV data file, with comma text preserved."""

    # Given: a TSV input and a TSV-shaped harmonized intermediate
    content = b"col_a\tcol_b\nalpha, beta\tgamma\n"
    file_id = await upload_content(app_client, content, "download.tsv", TEST_TSV_CONTENT_TYPE)
    meta = temp_storage.load(file_id)
    assert meta is not None
    harmonized_path = temp_storage.harmonized_path_for(file_id, meta.saved_path)
    with harmonized_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle, delimiter="\t")
        writer.writerows([["col_a", "col_b"], ["delta, epsilon", "gamma"]])
    store_test_completed_harmonization(temp_storage, file_id, harmonized_path)

    # When: the download endpoint is invoked
    response = await app_client.post("/stage-5/download", json={"file_id": file_id})

    # Then: the zip contains TSV output and values are tab-delimited
    assert response.status_code == 200
    with zipfile.ZipFile(BytesIO(response.content), "r") as archive:
        names = archive.namelist()
    assert any(name.endswith(".tsv") for name in names)
    assert not any(name.endswith(".csv") for name in names)
    output_rows = _read_downloaded_tabular(response.content, ".tsv", "\t")
    assert output_rows == [["col_a", "col_b"], ["delta, epsilon", "gamma"]]


async def test_stage5_download_xlsx_input_exports_xlsx_selected_sheet(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """An XLSX upload downloads XLSX output while updating only the selected sheet."""

    # Given: an XLSX input where Stage 1 selected the second sheet
    content = create_xlsx_content({
        "Keep": [["status"], ["unchanged"]],
        "Patients": [["col_a", "col_b"], ["alpha", "gamma"]],
    })
    file_id = await upload_content(app_client, content, "download.xlsx", TEST_XLSX_CONTENT_TYPE)
    analyze_response = await app_client.post(
        "/stage-1/analyze",
        json={
            "file_id": file_id,
            "data_model_key": TEST_TARGET_SCHEMA,
            "external_version_number": "11.0.4",
            "sheet_name": "Patients",
        },
    )
    assert analyze_response.status_code == 200
    meta = temp_storage.load(file_id)
    assert meta is not None
    harmonized_path = temp_storage.harmonized_path_for(file_id, meta.saved_path)
    harmonized_dataset = dataset_from_rows(
        headers=["col_a", "col_b"],
        rows=[["delta", "gamma"]],
        source_format=TabularFormat.XLSX,
        sheet_name="Patients",
    )
    write_tabular(harmonized_path, harmonized_dataset, template_path=meta.saved_path)
    store_test_completed_harmonization(temp_storage, file_id, harmonized_path)

    # When: the download endpoint is invoked
    response = await app_client.post("/stage-5/download", json={"file_id": file_id})

    # Then: the zip contains XLSX output, and non-selected sheets are preserved
    assert response.status_code == 200
    with zipfile.ZipFile(BytesIO(response.content), "r") as archive:
        names = archive.namelist()
    assert any(name.endswith(".xlsx") for name in names)
    assert not any(name.endswith(".csv") for name in names)
    assert _read_downloaded_xlsx(response.content, "Keep") == [["status"], ["unchanged"]]
    assert _read_downloaded_xlsx(response.content, "Patients") == [["col_a", "col_b"], ["delta", "gamma"]]


async def test_stage5_download_ignores_invalid_row_keys(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Overrides for out-of-range row keys do not alter the output."""

    # Given: an uploaded file with harmonized output and invalid overrides
    rows = [["col_a"], ["alpha"], ["beta"]]
    file_id = await upload_content(app_client, create_csv_content(rows), "invalid-rows.csv")
    meta = temp_storage.load(file_id)
    assert meta is not None
    create_harmonized_csv(temp_storage, file_id, meta.saved_path, {})
    create_manifest_for_file(temp_storage, file_id, meta.saved_path, {})
    await app_client.post(
        "/stage-4/overrides",
        json={
            "file_id": file_id,
            "overrides": {
                "99": {"col_0000": {"human_value": "gamma", "original_value": "alpha"}},
            },
            "review_state": review_state_payload(),
        },
        headers={"If-None-Match": "*"},
    )

    # When: the download endpoint is invoked
    response = await app_client.post("/stage-5/download", json={"file_id": file_id})

    # Then: output rows remain unchanged
    assert response.status_code == 200
    output_rows = _read_downloaded_csv(response.content)
    assert output_rows[0]["col_a"] == "alpha"
    assert output_rows[1]["col_a"] == "beta"


async def test_stage5_summary_zero_changes_when_terms_equal(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Summary counts zero AI changes when harmonized values equal originals."""

    # Given: an uploaded file with no changes in the manifest
    rows = [["col_a"], ["alpha"], ["beta"]]
    file_id = await upload_content(app_client, create_csv_content(rows), "summary.csv")
    meta = temp_storage.load(file_id)
    assert meta is not None
    create_harmonized_csv(temp_storage, file_id, meta.saved_path, {})
    create_manifest_for_file(temp_storage, file_id, meta.saved_path, {})

    # When: summary is requested
    response = await app_client.post("/stage-5/summary", json={"file_id": file_id})

    # Then: AI changes are zero
    assert response.status_code == 200
    summary = response.json()
    total_ai_changes = sum(column["ai_changes"] for column in summary["column_summaries"])
    assert total_ai_changes == 0
