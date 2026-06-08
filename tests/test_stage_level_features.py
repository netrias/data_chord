"""Stage-level feature tests for upload, mapping, harmonization, review, and download."""

from __future__ import annotations

import asyncio
import csv
import io
import json
import time
import zipfile
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from typing import cast
from unittest.mock import MagicMock

import pytest
from httpx import AsyncClient
from netrias_client import TabularFormat, dataset_from_rows, write_tabular
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import src.domain.dependencies as dependencies
import src.stage_3_harmonize.router as stage_three_router
from src.domain.cde import CDEInfo
from src.domain.data_model_cache import clear_session_cache, get_session_cache
from src.domain.harmonize import HarmonizeResult, HarmonizeStatus
from src.domain.manifest import ManifestPayload
from src.domain.storage import UploadStorage, WorkflowFile
from src.domain.workflow_state import WorkflowState
from src.domain.workflow_state_store import load_workflow_state
from src.stage_3_harmonize.job_state import StageThreeJobState
from tests.conftest import (
    TEST_TARGET_SCHEMA,
    TEST_TSV_CONTENT_TYPE,
    TEST_XLSX_CONTENT_TYPE,
    create_csv_content,
    create_harmonized_csv,
    create_manifest_for_file,
    create_test_manifest_parquet,
    create_xlsx_content,
    review_state_payload,
    upload_content,
)

pytestmark = pytest.mark.asyncio


def _load_workflow_state(file_id: str) -> WorkflowState | None:
    return load_workflow_state(
        dependencies.get_workflow_storage(),
        dependencies.get_user_context(),
        file_id,
    )


def _load_json_artifact(file_id: str, kind: WorkflowFile) -> object | None:
    stored = dependencies.get_workflow_storage().read_json(
        dependencies.get_user_context(),
        file_id,
        kind,
    )
    return stored.data if stored is not None else None


def _read_downloaded_csv(response_bytes: bytes) -> list[dict[str, str]]:
    with zipfile.ZipFile(BytesIO(response_bytes), "r") as zf:
        csv_name = next(name for name in zf.namelist() if name.endswith(".csv"))
        csv_content = zf.read(csv_name).decode("utf-8")
    return list(csv.DictReader(io.StringIO(csv_content)))


def _read_downloaded_tabular(response_bytes: bytes, suffix: str, delimiter: str) -> list[list[str]]:
    with zipfile.ZipFile(BytesIO(response_bytes), "r") as zf:
        data_name = next(name for name in zf.namelist() if name.endswith(suffix))
        content = zf.read(data_name).decode("utf-8")
    return list(csv.reader(io.StringIO(content), delimiter=delimiter))


def _read_downloaded_xlsx(response_bytes: bytes, sheet_name: str) -> list[list[str]]:
    with zipfile.ZipFile(BytesIO(response_bytes), "r") as zf:
        workbook_name = next(name for name in zf.namelist() if name.endswith(".xlsx"))
        workbook_bytes = BytesIO(zf.read(workbook_name))
    workbook = load_workbook(workbook_bytes, data_only=True)
    sheet = cast(Worksheet, workbook[sheet_name])
    return [[str(value) if value is not None else "" for value in row] for row in sheet.iter_rows(values_only=True)]


async def _wait_for_stage_three_job(app_client: AsyncClient, job_id: str, file_id: str) -> dict[str, object]:
    for _ in range(50):
        response = await app_client.get(f"/stage-3/jobs/{job_id}", params={"file_id": file_id})
        assert response.status_code == 200
        body = response.json()
        if body["status"] in {"succeeded", "failed"}:
            return cast(dict[str, object], body)
        await asyncio.sleep(0.02)
    raise AssertionError(f"Stage 3 job did not finish: {job_id}")


async def test_stage_three_job_state_requires_timezone_aware_start() -> None:
    # Given / When / Then: persisted job timing rejects ambiguous local datetimes
    with pytest.raises(ValueError):
        StageThreeJobState(
            polling_job_id="polling-job",
            job_id="job",
            file_id="file",
            status=HarmonizeStatus.QUEUED,
            detail="queued",
            next_stage_url="/stage-4",
            started_at=datetime(2026, 5, 21),
        )

    job = StageThreeJobState(
        polling_job_id="polling-job",
        job_id="job",
        file_id="file",
        status=HarmonizeStatus.QUEUED,
        detail="queued",
        next_stage_url="/stage-4",
        started_at=datetime(2026, 5, 21, tzinfo=UTC),
    )
    assert job.started_at.tzinfo is UTC


async def test_stage1_upload_persists_exact_bytes(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Upload stores the exact CSV bytes for later processing."""

    # Given: a CSV payload and no files in storage yet
    content = create_csv_content([["col_a"], ["alpha"], ["beta"]])
    assert list(temp_storage._data_dir.glob("*.csv")) == []

    # When: the file is uploaded
    file_id = await upload_content(app_client, content, "bytes.csv")

    # Then: stored metadata and file contents match the upload
    meta = temp_storage.load(file_id)
    assert meta is not None, "Expected stored metadata for uploaded file"
    assert meta.size_bytes == len(content), "Stored size does not match upload size"
    assert meta.saved_path.read_bytes() == content, "Stored bytes do not match uploaded bytes"


async def test_stage1_upload_preserves_other_session_cde_cache(app_client: AsyncClient) -> None:
    """Uploading a new file must not discard another active session's CDE cache."""

    first_file_id: str | None = None
    second_file_id: str | None = None

    try:
        # Given: one uploaded file has session-scoped CDE metadata
        first_file_id = await upload_content(app_client, create_csv_content([["col_a"], ["alpha"]]), "first.csv")
        first_cache = get_session_cache(first_file_id)
        assert not first_cache.has_cdes()

        first_cache.set_cdes(
            [
                CDEInfo(
                    cde_id=1,
                    cde_key="primary_diagnosis",
                    description="Primary Diagnosis",
                    version_label="1",
                )
            ],
            data_model_key=TEST_TARGET_SCHEMA,
            external_version_number="11.0.4",
        )
        assert first_cache.has_cdes()

        # When: another file is uploaded
        second_file_id = await upload_content(app_client, create_csv_content([["col_b"], ["beta"]]), "second.csv")

        # Then: the first file's cache remains available for Stage 2 and later
        assert second_file_id != first_file_id
        assert get_session_cache(first_file_id).get_cde_by_key("primary_diagnosis") is not None
    finally:
        if first_file_id is not None:
            clear_session_cache(first_file_id)
        if second_file_id is not None:
            clear_session_cache(second_file_id)


async def test_stage1_upload_rejects_mismatched_content_type(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Upload rejects non-CSV content types."""

    # Given: CSV bytes with an unsupported content type
    content = create_csv_content([["col_a"], ["alpha"]])
    assert list(temp_storage._data_dir.glob("*.csv")) == []

    # When: the file is uploaded with a mismatched content type
    response = await app_client.post(
        "/stage-1/upload",
        files={"file": ("bad.json", content, "application/json")},
    )

    # Then: upload is rejected with 415
    assert response.status_code == 415


async def test_stage1_analyze_rejects_invalid_utf8(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Analyze returns 400 for invalid UTF-8 payloads."""

    # Given: bytes that are not valid UTF-8
    content = b"\xff\xfe\xfa\xfb"
    file_id = await upload_content(app_client, content, "invalid.csv")
    assert temp_storage.load_manifest(file_id) is None

    # When: analyze is requested
    response = await app_client.post(
        "/stage-1/analyze",
        json={"file_id": file_id, "target_schema": TEST_TARGET_SCHEMA, "target_external_version_number": "11.0.4"},
    )

    # Then: bad request is returned
    assert response.status_code == 400


async def test_stage1_analyze_handles_quoted_commas(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Analyze treats quoted commas as part of the value."""

    # Given: a CSV containing quoted commas
    content = b'col_a\n"alpha, beta"\n'
    file_id = await upload_content(app_client, content, "quoted.csv")
    assert temp_storage.load_manifest(file_id) is None

    # When: analyze is requested
    response = await app_client.post(
        "/stage-1/analyze",
        json={"file_id": file_id, "target_schema": TEST_TARGET_SCHEMA, "target_external_version_number": "11.0.4"},
    )

    # Then: analyze succeeds and summarizes the column
    assert response.status_code == 200
    column = response.json()["columns"][0]
    assert column["column_name"] == "col_a"
    assert column["inferred_type"] == "text"


async def test_stage1_analyze_handles_ragged_rows(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Analyze fills missing values for ragged rows."""

    # Given: a CSV with missing values in some rows
    content = b"col_a,col_b\nalpha,beta\ncharlie,\n"
    file_id = await upload_content(app_client, content, "ragged.csv")
    assert temp_storage.load_manifest(file_id) is None

    # When: analyze is requested
    response = await app_client.post(
        "/stage-1/analyze",
        json={"file_id": file_id, "target_schema": TEST_TARGET_SCHEMA, "target_external_version_number": "11.0.4"},
    )

    # Then: the full-column confidence reflects the missing cell
    assert response.status_code == 200
    data = response.json()
    col_b = next(col for col in data["columns"] if col["column_name"] == "col_b")
    assert col_b["confidence_score"] == 0.5


async def test_stage1_analyze_accepts_duplicate_headers_with_distinct_column_keys(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Analyze preserves duplicate headers by assigning distinct column keys."""

    # Given: a CSV with duplicate header names
    content = b"col_a,col_a\nalpha,beta\n"
    file_id = await upload_content(app_client, content, "dupe.csv")
    assert temp_storage.load_manifest(file_id) is None

    # When: analyze is requested
    response = await app_client.post(
        "/stage-1/analyze",
        json={"file_id": file_id, "target_schema": TEST_TARGET_SCHEMA, "target_external_version_number": "11.0.4"},
    )

    # Then: both duplicate columns are present and independently addressable
    assert response.status_code == 200
    columns = response.json()["columns"]
    assert [column["column_name"] for column in columns] == ["col_a", "col_a"]
    assert [column["column_key"] for column in columns] == ["col_0000", "col_0001"]


async def test_stage1_analyze_accepts_blank_middle_header_by_column_position(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Analyze preserves blank display headers while keeping positional column keys."""

    # Given: a CSV with a blank middle header and no stored manifest yet
    content = b"col_a,,col_c\nalpha,beta,gamma\n"
    file_id = await upload_content(app_client, content, "blank-header.csv")
    assert temp_storage.load_manifest(file_id) is None

    # When: analyze is requested
    response = await app_client.post(
        "/stage-1/analyze",
        json={"file_id": file_id, "target_schema": TEST_TARGET_SCHEMA, "target_external_version_number": "11.0.4"},
    )

    # Then: the blank display name is allowed and the source column remains addressable by key
    assert response.status_code == 200
    columns = response.json()["columns"]
    assert [column["column_name"] for column in columns] == ["col_a", "", "col_c"]
    assert [column["column_key"] for column in columns] == ["col_0000", "col_0001", "col_0002"]
    assert [column["source_index"] for column in columns] == [0, 1, 2]


async def test_stage1_analyze_accepts_tsv(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Analyze treats tabs as column delimiters for TSV uploads."""

    # Given: a TSV with commas inside values and no manifest stored yet
    content = b"col_a\tcol_b\nalpha, beta\tgamma\n"
    file_id = await upload_content(app_client, content, "data.tsv", TEST_TSV_CONTENT_TYPE)
    assert temp_storage.load_manifest(file_id) is None

    # When: analyze is requested
    response = await app_client.post(
        "/stage-1/analyze",
        json={"file_id": file_id, "target_schema": TEST_TARGET_SCHEMA, "target_external_version_number": "11.0.4"},
    )

    # Then: tab-separated columns are parsed
    assert response.status_code == 200
    columns = response.json()["columns"]
    assert [column["column_name"] for column in columns] == ["col_a", "col_b"]


async def test_stage1_analyze_xlsx_defaults_to_first_sheet(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Analyze reads the first worksheet by default for XLSX uploads."""

    # Given: an XLSX workbook with distinct values on each sheet
    content = create_xlsx_content({
        "First": [["first_col"], ["first-value"]],
        "Second": [["second_col"], ["second-value"]],
    })
    file_id = await upload_content(app_client, content, "data.xlsx", TEST_XLSX_CONTENT_TYPE)
    assert temp_storage.load_manifest(file_id) is None

    # When: analyze is requested without a sheet override
    response = await app_client.post(
        "/stage-1/analyze",
        json={"file_id": file_id, "target_schema": TEST_TARGET_SCHEMA, "target_external_version_number": "11.0.4"},
    )

    # Then: analyzed columns come from the first sheet
    assert response.status_code == 200
    columns = response.json()["columns"]
    assert [column["column_name"] for column in columns] == ["first_col"]


async def test_stage1_analyze_xlsx_uses_selected_sheet(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Analyze uses the sheet selected in Stage 1 for XLSX uploads."""

    # Given: an XLSX workbook with duplicate headers on the second sheet
    content = create_xlsx_content({
        "First": [["ignored"], ["nope"]],
        "Patients": [["col_a", "col_a"], ["alpha", "beta"]],
    })
    file_id = await upload_content(app_client, content, "data.xlsx", TEST_XLSX_CONTENT_TYPE)
    assert temp_storage.load_manifest(file_id) is None

    # When: analyze is requested for the second sheet
    response = await app_client.post(
        "/stage-1/analyze",
        json={
            "file_id": file_id,
            "target_schema": TEST_TARGET_SCHEMA,
            "target_external_version_number": "11.0.4",
            "sheet_name": "Patients",
        },
    )

    # Then: selected-sheet columns are parsed without collapsing duplicate headers
    assert response.status_code == 200
    columns = response.json()["columns"]
    assert [column["column_name"] for column in columns] == ["col_a", "col_a"]
    assert [column["column_key"] for column in columns] == ["col_0000", "col_0001"]
    meta = temp_storage.load(file_id)
    assert meta is not None
    assert meta.selected_sheet == "Patients"


async def test_stage1_analyze_bom_and_non_bom_match_headers(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """BOM and non-BOM files produce identical headers."""

    # Given: BOM and non-BOM CSVs with the same headers
    bom_content = "\ufeffcol_a,col_b\nalpha,beta\n".encode()
    non_bom_content = b"col_a,col_b\nalpha,beta\n"
    bom_file_id = await upload_content(app_client, bom_content, "bom.csv")
    non_bom_file_id = await upload_content(app_client, non_bom_content, "plain.csv")
    assert temp_storage.load_manifest(bom_file_id) is None
    assert temp_storage.load_manifest(non_bom_file_id) is None

    # When: analyze is requested for both
    bom_response = await app_client.post(
        "/stage-1/analyze",
        json={
            "file_id": bom_file_id,
            "target_schema": TEST_TARGET_SCHEMA,
            "target_external_version_number": "11.0.4",
        },
    )
    non_bom_response = await app_client.post(
        "/stage-1/analyze",
        json={
            "file_id": non_bom_file_id,
            "target_schema": TEST_TARGET_SCHEMA,
            "target_external_version_number": "11.0.4",
        },
    )

    # Then: the headers are identical
    assert bom_response.status_code == 200
    assert non_bom_response.status_code == 200
    bom_headers = [col["column_name"] for col in bom_response.json()["columns"]]
    non_bom_headers = [col["column_name"] for col in non_bom_response.json()["columns"]]
    assert bom_headers == non_bom_headers

async def test_stage1_analyze_handles_bom_headers(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Analyze strips BOM headers so column names are correct."""

    # Given: a BOM-prefixed CSV and no manifest stored yet
    content = "\ufeffrecord_id,col_a\nRID-1,Foo\n".encode()
    file_id = await upload_content(app_client, content, "bom.csv")
    assert temp_storage.load_manifest(file_id) is None

    # When: the file is analyzed
    response = await app_client.post(
        "/stage-1/analyze",
        json={"file_id": file_id, "target_schema": TEST_TARGET_SCHEMA, "target_external_version_number": "11.0.4"},
    )

    # Then: column names do not include BOM characters
    assert response.status_code == 200
    columns = response.json()["columns"]
    assert columns[0]["column_name"] == "record_id"


async def test_stage1_analyze_is_idempotent(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Repeated analysis does not change the stored manifest."""

    # Given: an uploaded CSV with no manifest yet
    content = create_csv_content([["col_a"], ["alpha"], ["beta"]])
    file_id = await upload_content(app_client, content, "idempotent.csv")
    assert temp_storage.load_manifest(file_id) is None

    # When: the file is analyzed twice
    response_one = await app_client.post(
        "/stage-1/analyze",
        json={"file_id": file_id, "target_schema": TEST_TARGET_SCHEMA, "target_external_version_number": "11.0.4"},
    )
    response_two = await app_client.post(
        "/stage-1/analyze",
        json={"file_id": file_id, "target_schema": TEST_TARGET_SCHEMA, "target_external_version_number": "11.0.4"},
    )

    # Then: manifest and API outputs remain stable
    assert response_one.status_code == 200
    assert response_two.status_code == 200
    manifest_one = temp_storage.load_manifest(file_id)
    manifest_two = temp_storage.load_manifest(file_id)
    assert manifest_one == manifest_two, "Manifest changed between analyses"


async def test_stage1_analyze_uses_selected_version_and_primes_reference_cache(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
    mock_netrias_client: MagicMock,
) -> None:
    """Analyze passes the selected model version and warms CDE/PV cache."""

    # Given: an uploaded CSV and an empty session cache
    content = create_csv_content([["diagnosis"], ["Lung"], ["Breast"]])
    file_id = await upload_content(app_client, content, "versioned.csv")
    cache = get_session_cache(file_id)
    assert not cache.has_cdes()
    assert not cache.has_any_pvs()

    # When: analysis is requested for GC external version 11.0.4
    response = await app_client.post(
        "/stage-1/analyze",
        json={
            "file_id": file_id,
            "target_schema": "gc",
            "target_external_version_number": "11.0.4",
        },
    )

    # Then: discovery and the session cache use the selected external version
    assert response.status_code == 200
    assert response.json()["target_external_version_number"] == "11.0.4"
    assert mock_netrias_client.discover_mapping_from_tabular.call_args.kwargs["target_version"] == "11.0.4"
    data_model_version = cache.get_data_model_version()
    assert data_model_version is not None
    assert data_model_version.data_model_key == "gc"
    assert data_model_version.external_version_number == "11.0.4"
    assert cache.has_cdes()
    assert cache.has_any_pvs()


async def test_stage1_analyze_persists_selected_data_model_version(
    app_client: AsyncClient,
) -> None:
    """Analyze saves the selected model/version as durable workflow state."""

    # Given: an uploaded CSV with no workflow selection saved yet
    file_id = await upload_content(app_client, create_csv_content([["diagnosis"], ["Lung"]]), "selection.csv")
    assert _load_workflow_state(file_id) is None

    # When: analysis is requested for a specific external model version
    response = await app_client.post(
        "/stage-1/analyze",
        json={
            "file_id": file_id,
            "target_schema": "gc",
            "target_external_version_number": "11.0.4",
        },
    )

    # Then: the selected model/version is available from durable workflow state
    assert response.status_code == 200
    state = _load_workflow_state(file_id)
    assert state is not None
    assert state.file_id == file_id
    assert state.data_model_version.data_model_key == "gc"
    assert state.data_model_version.external_version_number == "11.0.4"


async def test_stage2_mapping_page_recovers_selected_model_from_workflow_state(
    app_client: AsyncClient,
    mock_netrias_client: MagicMock,
) -> None:
    """Stage 2 can reload after cache loss using the durable selected model/version."""

    # Given: analysis saved GC external version 11.0.4, then the in-memory CDE cache was lost
    file_id = await upload_content(app_client, create_csv_content([["diagnosis"], ["Lung"]]), "stage2-selection.csv")
    analyze_response = await app_client.post(
        "/stage-1/analyze",
        json={
            "file_id": file_id,
            "target_schema": "gc",
            "target_external_version_number": "11.0.4",
        },
    )
    assert analyze_response.status_code == 200
    clear_session_cache(file_id)
    mock_netrias_client.list_cdes.reset_mock()

    # When: Stage 2 is loaded with only the file id in the URL
    response = await app_client.get(f"/stage-2?file_id={file_id}")

    # Then: the page uses the stored selection to rebuild CDE options
    assert response.status_code == 200
    assert 'targetSchema: "gc"' in response.text
    assert 'targetExternalVersionNumber: "11.0.4"' in response.text
    assert "stage_2_mappings.js?v=" in response.text
    mock_netrias_client.list_cdes.assert_called_with(
        model_key="gc",
        version="11.0.4",
        include_description=True,
    )


async def test_static_assets_require_browser_revalidation(app_client: AsyncClient) -> None:
    """Deployed browsers must not keep stale stage JavaScript across releases."""

    response = await app_client.get("/assets/stage-2/stage_2_mappings.js?v=test-release")

    assert response.status_code == 200
    assert "must-revalidate" in response.headers["Cache-Control"]


async def test_stage3_harmonize_prefers_stored_selection_over_stale_request(
    app_client: AsyncClient,
) -> None:
    """The durable selected model/version is backend truth during harmonization."""

    class StubHarmonizer:
        def __init__(self) -> None:
            self.received_data_model_key = None
            self.received_target_version = None

        def run(  # type: ignore[no-untyped-def]
            self,
            *,
            file_path,
            data_model_key,
            column_overrides,
            column_renames,
            cache,
            external_version_number,
            manifest,
            output_path,
            sheet_name,
        ):
            self.received_data_model_key = data_model_key
            self.received_target_version = external_version_number
            return HarmonizeResult(job_id="job-selection", status=HarmonizeStatus.SUCCEEDED, detail="ok")

    # Given: analysis saved GC external version 11.0.4, but the browser later sends stale request selection
    file_id = await upload_content(app_client, create_csv_content([["diagnosis"], ["Lung"]]), "stage3-selection.csv")
    analyze_response = await app_client.post(
        "/stage-1/analyze",
        json={
            "file_id": file_id,
            "target_schema": "gc",
            "target_external_version_number": "11.0.4",
        },
    )
    assert analyze_response.status_code == 200
    stub = StubHarmonizer()

    # When: harmonization is triggered with stale model/version fields
    import unittest.mock

    with unittest.mock.patch("src.stage_3_harmonize.router.get_harmonize_service", return_value=stub):
        response = await app_client.post(
            "/stage-3/harmonize",
            json={
                "file_id": file_id,
                "target_schema": "stale-model",
                "target_external_version_number": "99.0.0",
                "manual_overrides": {},
            },
        )

    # Then: the harmonization service receives the stored selection instead
    assert response.status_code == 200
    assert stub.received_data_model_key == "gc"
    assert stub.received_target_version == "11.0.4"


async def test_stage3_harmonize_returns_queued_while_long_job_finishes(
    app_client: AsyncClient,
) -> None:
    """Stage 3 does not keep the browser request open for slow harmonization jobs."""

    class SlowStubHarmonizer:
        def run(  # type: ignore[no-untyped-def]
            self,
            *,
            file_path,
            data_model_key,
            column_overrides,
            column_renames,
            cache,
            external_version_number,
            manifest,
            output_path,
            sheet_name,
        ):
            time.sleep(0.5)
            return HarmonizeResult(job_id="job-slow", status=HarmonizeStatus.SUCCEEDED, detail="ok")

    # Given: an analyzed upload has no completed Stage 3 job yet
    file_id = await upload_content(app_client, create_csv_content([["diagnosis"], ["Lung"]]), "slow-stage3.csv")
    analyze_response = await app_client.post(
        "/stage-1/analyze",
        json={"file_id": file_id, "target_schema": "gc", "target_external_version_number": "11.0.4"},
    )
    assert analyze_response.status_code == 200

    # When: harmonization is triggered and the harmonizer is still running
    import unittest.mock

    with unittest.mock.patch("src.stage_3_harmonize.router.get_harmonize_service", return_value=SlowStubHarmonizer()):
        response = await app_client.post(
            "/stage-3/harmonize",
            json={
                "file_id": file_id,
                "target_schema": "gc",
                "target_external_version_number": "11.0.4",
                "manual_overrides": {},
            },
        )

        # Then: the browser gets a queued job promptly and can poll it to completion
        assert response.status_code == 200
        body = response.json()
        assert body["status"] == "queued"
        assert isinstance(body["elapsed_seconds"], int)
        queued_next_stage_url = body["next_stage_url"]
        assert isinstance(queued_next_stage_url, str)
        assert "status=queued" in queued_next_stage_url
        assert body["manifest_summary"] is None

        finished = await _wait_for_stage_three_job(app_client, body["job_id"], file_id)

    assert finished["status"] == "succeeded"
    assert finished["job_id"] == "job-slow"
    assert isinstance(finished["elapsed_seconds"], int)
    finished_next_stage_url = finished["next_stage_url"]
    assert isinstance(finished_next_stage_url, str)
    assert "status=succeeded" in finished_next_stage_url


async def test_stage3_job_status_recovers_from_durable_state_after_cache_loss(
    app_client: AsyncClient,
) -> None:
    """Stage 3 polling can recover after the process-local job cache is gone."""

    class SlowStubHarmonizer:
        def run(  # type: ignore[no-untyped-def]
            self,
            *,
            file_path,
            data_model_key,
            column_overrides,
            column_renames,
            cache,
            external_version_number,
            manifest,
            output_path,
            sheet_name,
        ):
            time.sleep(0.5)
            return HarmonizeResult(job_id="job-durable", status=HarmonizeStatus.SUCCEEDED, detail="ok")

    # Given: a slow Stage 3 job has been accepted and later completed
    file_id = await upload_content(app_client, create_csv_content([["diagnosis"], ["Lung"]]), "durable-stage3.csv")
    analyze_response = await app_client.post(
        "/stage-1/analyze",
        json={"file_id": file_id, "target_schema": "gc", "target_external_version_number": "11.0.4"},
    )
    assert analyze_response.status_code == 200
    assert _load_json_artifact(file_id, WorkflowFile.STAGE_THREE_JOB) is None

    import unittest.mock

    with unittest.mock.patch("src.stage_3_harmonize.router.get_harmonize_service", return_value=SlowStubHarmonizer()):
        response = await app_client.post(
            "/stage-3/harmonize",
            json={
                "file_id": file_id,
                "target_schema": "gc",
                "target_external_version_number": "11.0.4",
                "manual_overrides": {},
            },
        )
        assert response.status_code == 200
        accepted_job_id = response.json()["job_id"]
        assert _load_json_artifact(file_id, WorkflowFile.STAGE_THREE_JOB) is not None

        finished = await _wait_for_stage_three_job(app_client, accepted_job_id, file_id)

    # When: the in-memory job cache disappears, like it would after a deploy
    stage_three_router._stage_three_jobs.clear()
    recovered_response = await app_client.get(
        f"/stage-3/jobs/{accepted_job_id}",
        params={"file_id": file_id},
    )

    # Then: the poll endpoint falls back to workflow storage
    assert finished["status"] == "succeeded"
    assert recovered_response.status_code == 200
    recovered = recovered_response.json()
    assert recovered["status"] == "succeeded"
    assert recovered["job_id"] == "job-durable"
    assert isinstance(recovered["elapsed_seconds"], int)


async def test_stage2_saves_confirmed_mapping_choices_to_workflow_state(
    app_client: AsyncClient,
) -> None:
    """Stage 2 persists user mapping and rename choices before Stage 3."""

    # Given: analysis has created workflow state for an uploaded file
    file_id = await upload_content(app_client, create_csv_content([["diagnosis"], ["Lung"]]), "mapping-choices.csv")
    analyze_response = await app_client.post(
        "/stage-1/analyze",
        json={"file_id": file_id, "target_schema": "gc", "target_external_version_number": "11.0.4"},
    )
    assert analyze_response.status_code == 200
    state = _load_workflow_state(file_id)
    assert state is not None
    assert state.mapping_choices is None

    # When: Stage 2 saves confirmed choices
    response = await app_client.post(
        "/stage-2/choices",
        json={
            "file_id": file_id,
            "manual_overrides": {"col_0000": "primary_diagnosis", "col_0001": None},
            "column_renames": {"col_0000": "Primary Diagnosis"},
        },
    )

    # Then: the choices are durable workflow state
    assert response.status_code == 200
    updated = _load_workflow_state(file_id)
    assert updated is not None
    assert updated.mapping_choices is not None
    assert updated.mapping_choices.column_overrides.to_strings() == {
        "col_0000": "primary_diagnosis",
        "col_0001": None,
    }
    assert updated.mapping_choices.column_renames.to_strings() == {"col_0000": "Primary Diagnosis"}


async def test_stage2_save_mapping_choices_requires_workflow_state(
    app_client: AsyncClient,
) -> None:
    """Stage 2 reports a clear error when choices are saved before analysis."""

    # Given: a file was uploaded, but Stage 1 analysis has not created workflow state
    file_id = await upload_content(app_client, create_csv_content([["diagnosis"], ["Lung"]]), "no-state.csv")
    assert _load_workflow_state(file_id) is None

    # When: Stage 2 tries to persist confirmed choices
    response = await app_client.post(
        "/stage-2/choices",
        json={
            "file_id": file_id,
            "manual_overrides": {"col_0000": "primary_diagnosis"},
            "column_renames": {"col_0000": "Primary Diagnosis"},
        },
    )

    # Then: the endpoint preserves the existing missing-workflow-state contract
    assert response.status_code == 404
    assert response.json()["detail"] == "Workflow state not found. Please rerun analysis."


async def test_stage3_harmonize_prefers_stored_mapping_choices_over_stale_request(
    app_client: AsyncClient,
) -> None:
    """The confirmed Stage 2 choices are backend truth during harmonization."""

    class StubHarmonizer:
        def __init__(self) -> None:
            self.received_overrides = None
            self.received_renames = None

        def run(  # type: ignore[no-untyped-def]
            self,
            *,
            file_path,
            data_model_key,
            column_overrides,
            column_renames,
            cache,
            external_version_number,
            manifest,
            output_path,
            sheet_name,
        ):
            self.received_overrides = column_overrides.to_strings()
            self.received_renames = column_renames.to_strings()
            return HarmonizeResult(job_id="job-choices", status=HarmonizeStatus.SUCCEEDED, detail="ok")

    # Given: Stage 2 saved confirmed mapping choices
    file_id = await upload_content(app_client, create_csv_content([["diagnosis"], ["Lung"]]), "stage3-choices.csv")
    analyze_response = await app_client.post(
        "/stage-1/analyze",
        json={"file_id": file_id, "target_schema": "gc", "target_external_version_number": "11.0.4"},
    )
    assert analyze_response.status_code == 200
    choices_response = await app_client.post(
        "/stage-2/choices",
        json={
            "file_id": file_id,
            "manual_overrides": {"col_0000": "primary_diagnosis", "col_0001": None},
            "column_renames": {"col_0000": "Primary Diagnosis"},
        },
    )
    assert choices_response.status_code == 200
    stub = StubHarmonizer()

    # When: the browser sends stale choices in the harmonize request
    import unittest.mock

    with unittest.mock.patch("src.stage_3_harmonize.router.get_harmonize_service", return_value=stub):
        response = await app_client.post(
            "/stage-3/harmonize",
            json={
                "file_id": file_id,
                "target_schema": "stale-model",
                "target_external_version_number": "99.0.0",
                "manual_overrides": {"col_0000": "therapeutic_agents"},
                "column_renames": {"col_0000": "Stale Name"},
            },
        )

    # Then: Stage 3 uses the confirmed choices from workflow state
    assert response.status_code == 200
    assert stub.received_overrides == {"col_0000": "primary_diagnosis", "col_0001": None}
    assert stub.received_renames == {"col_0000": "Primary Diagnosis"}


async def test_stage3_applies_confirmed_column_renames_to_download(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
    tmp_path: Path,
) -> None:
    """Confirmed Stage 2 rename choices become final download headers."""

    class StubHarmonizer:
        def __init__(self) -> None:
            self.received_renames: dict[str, str] | None = None

        def run(  # type: ignore[no-untyped-def]
            self,
            *,
            file_path,
            data_model_key,
            column_overrides,
            column_renames,
            cache,
            external_version_number,
            manifest,
            output_path,
            sheet_name,
        ):
            with output_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.writer(handle)
                writer.writerows([["diagnosis"], ["Lung Cancer"]])
            assert output_path.read_text(encoding="utf-8").splitlines()[0] == "diagnosis"
            manifest_path = tmp_path / "rename-manifest.parquet"
            create_test_manifest_parquet(
                manifest_path,
                [
                    {
                        "job_id": "job-renamed-download",
                        "column_id": 0,
                        "column_name": "diagnosis",
                        "to_harmonize": "Lung",
                        "top_harmonization": "Lung Cancer",
                        "ontology_id": None,
                        "top_harmonizations": ["Lung Cancer"],
                        "confidence_score": 0.95,
                        "error": None,
                        "row_indices": [0],
                        "manual_overrides": [],
                    }
                ],
            )
            return HarmonizeResult(
                job_id="job-renamed-download",
                status=HarmonizeStatus.SUCCEEDED,
                detail="ok",
                manifest_path=manifest_path,
            )

    # Given: Stage 2 saved a confirmed output rename for the analyzed column
    file_id = await upload_content(app_client, create_csv_content([["diagnosis"], ["Lung"]]), "renamed-flow.csv")
    analyze_response = await app_client.post(
        "/stage-1/analyze",
        json={"file_id": file_id, "target_schema": TEST_TARGET_SCHEMA, "target_external_version_number": "11.0.4"},
    )
    assert analyze_response.status_code == 200
    choices_response = await app_client.post(
        "/stage-2/choices",
        json={
            "file_id": file_id,
            "manual_overrides": {},
            "column_renames": {"col_0000": "Primary Diagnosis"},
        },
    )
    assert choices_response.status_code == 200

    # When: Stage 3 runs and Stage 5 downloads the final dataset
    import unittest.mock

    with unittest.mock.patch("src.stage_3_harmonize.router.get_harmonize_service", return_value=StubHarmonizer()):
        harmonize_response = await app_client.post(
            "/stage-3/harmonize",
            json={
                "file_id": file_id,
                "target_schema": TEST_TARGET_SCHEMA,
                "target_external_version_number": "11.0.4",
                "manual_overrides": {},
                "column_renames": {},
            },
        )
    assert harmonize_response.status_code == 200

    rows_response = await app_client.post("/stage-4/rows", json={"file_id": file_id})
    summary_response = await app_client.post("/stage-5/summary", json={"file_id": file_id})
    download_response = await app_client.post("/stage-5/download", json={"file_id": file_id})

    # Then: review, summary, and download use the confirmed output header, not the source header
    assert rows_response.status_code == 200
    assert rows_response.json()["columns"][0]["columnLabel"] == "Primary Diagnosis"
    assert summary_response.status_code == 200
    assert summary_response.json()["column_summaries"][0]["column"] == "Primary Diagnosis"
    assert download_response.status_code == 200
    output_rows = _read_downloaded_csv(download_response.content)
    assert list(output_rows[0]) == ["Primary Diagnosis"]
    assert output_rows[0]["Primary Diagnosis"] == "Lung Cancer"


async def test_stage3_column_renames_propagate_when_output_name_matches_existing_header(
    app_client: AsyncClient,
    tmp_path: Path,
) -> None:
    """Rename choices are keyed by column identity, even when output names repeat."""

    class StubHarmonizer:
        def run(  # type: ignore[no-untyped-def]
            self,
            *,
            file_path,
            data_model_key,
            column_overrides,
            column_renames,
            cache,
            external_version_number,
            manifest,
            output_path,
            sheet_name,
        ):
            self.received_renames = column_renames.to_strings()
            with output_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.writer(handle)
                writer.writerows([["disease_type", "disease_type"], ["stale", ""]])
            actual_output_path = output_path.with_name(f"{output_path.stem}.v2{output_path.suffix}")
            with actual_output_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.writer(handle)
                writer.writerows([["diagnosis", "disease_type"], ["Lung Cancer", ""]])
            manifest_path = tmp_path / "repeat-output-name-manifest.parquet"
            create_test_manifest_parquet(
                manifest_path,
                [
                    {
                        "job_id": "job-repeat-output-name",
                        "column_id": 0,
                        "column_name": "diagnosis",
                        "to_harmonize": "Lung",
                        "top_harmonization": "Lung Cancer",
                        "ontology_id": None,
                        "top_harmonizations": ["Lung Cancer"],
                        "confidence_score": 0.95,
                        "error": None,
                        "row_indices": [0],
                        "manual_overrides": [],
                    }
                ],
            )
            return HarmonizeResult(
                job_id="job-repeat-output-name",
                status=HarmonizeStatus.SUCCEEDED,
                detail="ok",
                manifest_path=manifest_path,
                output_path=actual_output_path,
            )

    # Given: the uploaded file already contains the target standard column name
    stub = StubHarmonizer()
    file_id = await upload_content(
        app_client,
        create_csv_content([["diagnosis", "disease_type"], ["Lung", ""]]),
        "repeat-output-name-flow.csv",
    )
    analyze_response = await app_client.post(
        "/stage-1/analyze",
        json={"file_id": file_id, "target_schema": TEST_TARGET_SCHEMA, "target_external_version_number": "11.0.4"},
    )
    assert analyze_response.status_code == 200
    choices_response = await app_client.post(
        "/stage-2/choices",
        json={
            "file_id": file_id,
            "manual_overrides": {},
            "column_renames": {"col_0000": "disease_type"},
        },
    )
    assert choices_response.status_code == 200

    # When: Stage 3 runs and Stage 5 downloads the final dataset
    import unittest.mock

    with unittest.mock.patch("src.stage_3_harmonize.router.get_harmonize_service", return_value=stub):
        harmonize_response = await app_client.post(
            "/stage-3/harmonize",
            json={
                "file_id": file_id,
                "target_schema": TEST_TARGET_SCHEMA,
                "target_external_version_number": "11.0.4",
                "manual_overrides": {},
                "column_renames": {},
            },
        )
    assert harmonize_response.status_code == 200

    rows_response = await app_client.post("/stage-4/rows", json={"file_id": file_id})
    summary_response = await app_client.post("/stage-5/summary", json={"file_id": file_id})
    download_response = await app_client.post("/stage-5/download", json={"file_id": file_id})
    cde_mapping = _load_json_artifact(file_id, WorkflowFile.CDE_MAPPING)

    # Then: the selected output name propagates for col_0000 without merging it into col_0001
    assert rows_response.status_code == 200
    assert stub.received_renames == {"col_0000": "disease_type"}
    assert [column["columnLabel"] for column in rows_response.json()["columns"]] == ["disease_type"]
    assert summary_response.status_code == 200
    assert [summary["column"] for summary in summary_response.json()["column_summaries"]] == ["disease_type"]
    assert download_response.status_code == 200
    output_rows = _read_downloaded_tabular(download_response.content, ".csv", ",")
    output_headers = output_rows[0]
    assert output_headers == ["disease_type", "disease_type"]
    assert len(output_headers) == 2
    assert len(set(output_headers)) == 1
    assert output_rows[1] == ["Lung Cancer", ""]
    assert isinstance(cde_mapping, dict)
    mappings = {mapping["column_key"]: mapping for mapping in cde_mapping["mappings"]}
    assert mappings["col_0000"]["output_column_name"] == "disease_type"
    assert mappings["col_0001"]["output_column_name"] == "disease_type"


async def test_stage3_persists_cde_mapping_download_artifact(
    app_client: AsyncClient,
) -> None:
    """Harmonize saves the column-to-CDE mapping plan for the download bundle."""

    # Given: an uploaded CSV and a manifest with two mapped columns
    file_id = await upload_content(
        app_client,
        create_csv_content([["diagnosis", "drug"], ["Lung", "Agent A"]]),
        "mapping-plan.csv",
    )
    cache = get_session_cache(file_id)
    cache.set_cdes(
        [
            CDEInfo(cde_id=2, cde_key="primary_diagnosis", description="Primary Diagnosis", version_label="1"),
            CDEInfo(cde_id=1, cde_key="therapeutic_agents", description="Therapeutic Agents", version_label="1"),
        ],
        data_model_key=TEST_TARGET_SCHEMA,
        external_version_number="11.0.4",
    )
    manifest: ManifestPayload = {
        "column_mappings": {
            "col_0000": {"column_name": "diagnosis", "cde_key": "primary_diagnosis", "cde_id": 2},
            "col_0001": {"column_name": "drug", "cde_key": "therapeutic_agents", "cde_id": 1},
        }
    }

    # When: the user overrides and renames the second column
    response = await app_client.post(
        "/stage-3/harmonize",
        json={
            "file_id": file_id,
            "target_schema": TEST_TARGET_SCHEMA,
            "target_external_version_number": "11.0.4",
            "manual_overrides": {"col_0001": "primary_diagnosis"},
            "column_renames": {"col_0001": "Treatment Diagnosis"},
            "manifest": manifest,
        },
    )

    # Then: a mapping artifact records AI mappings, user overrides, and output names by column key
    assert response.status_code == 200
    document = _load_json_artifact(file_id, WorkflowFile.CDE_MAPPING)
    assert isinstance(document, dict)
    mappings = {entry["column_key"]: entry for entry in document["mappings"]}
    assert mappings["col_0000"]["mapping_source"] == "ai"
    assert mappings["col_0000"]["cde_key"] == "primary_diagnosis"
    assert mappings["col_0001"]["mapping_source"] == "user_override"
    assert mappings["col_0001"]["source_column_name"] == "drug"
    assert mappings["col_0001"]["output_column_name"] == "Treatment Diagnosis"
    assert mappings["col_0001"]["cde_description"] == "Primary Diagnosis"


async def test_stage2_mapping_page_renders_manual_options(
    app_client: AsyncClient,
) -> None:
    """Stage 2 mapping page exposes CDE labels for manual mapping."""

    # Given: CDE cache is pre-populated for the file
    file_id = "deadbeefdeadbeefdeadbeefdeadbeef"
    cache = get_session_cache(file_id)
    cache.set_cdes(
        [CDEInfo(cde_id=2, cde_key="primary_diagnosis", description=None, version_label="v1")],
        data_model_key="test-data-model",
        external_version_number="11.0.4",
    )

    # When: the mapping page is requested with schema query param
    response = await app_client.get(
        f"/stage-2?file_id={file_id}&schema=test-data-model&external_version_number=11.0.4"
    )

    # Then: the page renders and includes CDE labels
    assert response.status_code == 200
    assert "primary_diagnosis" in response.text


async def test_stage2_mapping_page_includes_default_schema(
    app_client: AsyncClient,
) -> None:
    """Stage 2 mapping page renders the schema from query param."""

    # Given: the mapping page has not been loaded yet
    # When: the mapping page is requested with schema query param
    response = await app_client.get("/stage-2?schema=test-data-model")

    # Then: the schema is embedded for client-side use
    assert response.status_code == 200
    assert 'targetSchema: "test-data-model"' in response.text


async def test_stage3_harmonize_uses_stored_manifest_when_payload_missing(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Harmonize falls back to the stored manifest if request payload omits it."""

    class StubHarmonizer:
        def __init__(self) -> None:
            self.received_manifest = None

        def run(  # type: ignore[no-untyped-def]
            self,
            *,
            file_path,
            data_model_key,
            column_overrides,
            column_renames,
            cache,
            external_version_number,
            manifest,
            output_path,
            sheet_name,
        ):
            self.received_manifest = manifest
            return HarmonizeResult(job_id="job-1", status=HarmonizeStatus.SUCCEEDED, detail="ok")

    # Given: an uploaded file with a stored manifest
    file_id = await upload_content(app_client, create_csv_content([["col_a"], ["alpha"]]), "manifest.csv")
    stored_manifest: ManifestPayload = {"column_mappings": {"col_a": {"cde_key": "primary_diagnosis", "cde_id": 2}}}
    temp_storage.save_manifest(file_id, stored_manifest)
    stub = StubHarmonizer()
    assert stub.received_manifest is None

    # When: harmonize is triggered without a manifest payload
    import unittest.mock

    with unittest.mock.patch("src.stage_3_harmonize.router.get_harmonize_service", return_value=stub):
        response = await app_client.post(
            "/stage-3/harmonize",
            json={
                "file_id": file_id,
                "target_schema": TEST_TARGET_SCHEMA,
                "target_external_version_number": "11.0.4",
                "manual_overrides": {},
            },
        )

    # Then: the stored manifest is used
    assert response.status_code == 200
    assert stub.received_manifest == {
        "column_mappings": {
            "col_a": {
                "column_name": "col_a",
                "cde_key": "primary_diagnosis",
                "cde_id": 2,
                "harmonization": "harmonizable",
                "alternatives": [],
            }
        }
    }


async def test_stage3_harmonize_prefers_stored_manifest_over_payload_manifest(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """The stored manifest is the backend source of truth when both copies exist."""

    class StubHarmonizer:
        def __init__(self) -> None:
            self.received_manifest = None
            self.received_target_version = None

        def run(  # type: ignore[no-untyped-def]
            self,
            *,
            file_path,
            data_model_key,
            column_overrides,
            column_renames,
            cache,
            external_version_number,
            manifest,
            output_path,
            sheet_name,
        ):
            self.received_manifest = manifest
            self.received_target_version = external_version_number
            return HarmonizeResult(job_id="job-2", status=HarmonizeStatus.SUCCEEDED, detail="ok")

    # Given: an uploaded file with a stored manifest and a stale request manifest
    file_id = await upload_content(app_client, create_csv_content([["col_a"], ["alpha"]]), "payload.csv")
    temp_storage.save_manifest(
        file_id,
        {"column_mappings": {"col_a": {"cde_key": "primary_diagnosis", "cde_id": 2}}},
    )
    payload_manifest: ManifestPayload = {"column_mappings": {"col_a": {"cde_key": "morphology", "cde_id": 3}}}
    stub = StubHarmonizer()

    # When: harmonize is triggered with the stale manifest payload
    import unittest.mock

    with unittest.mock.patch("src.stage_3_harmonize.router.get_harmonize_service", return_value=stub):
        response = await app_client.post(
            "/stage-3/harmonize",
            json={
                "file_id": file_id,
                "target_schema": TEST_TARGET_SCHEMA,
                "target_external_version_number": "11.0.4",
                "manual_overrides": {},
                "manifest": payload_manifest,
            },
        )

    # Then: the stored manifest is used instead of the stale request copy
    assert response.status_code == 200
    assert stub.received_manifest == {
        "column_mappings": {
            "col_a": {
                "column_name": "col_a",
                "cde_key": "primary_diagnosis",
                "cde_id": 2,
                "harmonization": "harmonizable",
                "alternatives": [],
            }
        }
    }
    assert stub.received_target_version == "11.0.4"


async def test_stage5_download_matches_harmonized_when_no_overrides(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Download returns the harmonized dataset when no overrides exist."""

    # Given: an uploaded file with harmonized output and no overrides
    rows = [["col_a"], ["alpha"], ["beta"]]
    file_id = await upload_content(app_client, create_csv_content(rows), "download.csv")
    meta = temp_storage.load(file_id)
    assert meta is not None
    create_harmonized_csv(temp_storage, file_id, meta.saved_path, {1: {"col_a": "gamma"}})
    create_manifest_for_file(temp_storage, file_id, meta.saved_path, {1: {"col_a": "gamma"}})

    # When: the download endpoint is invoked
    response = await app_client.post("/stage-5/download", json={"file_id": file_id})

    # Then: the CSV reflects harmonized values
    assert response.status_code == 200
    output_rows = _read_downloaded_csv(response.content)
    assert output_rows[1]["col_a"] == "gamma"


async def test_upload_storage_loads_managed_harmonized_output_path(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """UploadStorage owns harmonized output lookup by file id."""

    # Given: an uploaded file with no harmonized output yet
    file_id = await upload_content(app_client, create_csv_content([["col_a"], ["alpha"]]), "managed.csv")
    assert temp_storage.load_harmonized_path(file_id) is None

    # When: harmonized output is written to the managed storage location
    meta = temp_storage.load(file_id)
    assert meta is not None
    harmonized_path = temp_storage.harmonized_path_for(file_id, meta.saved_path)
    harmonized_path.write_text("col_a\nbeta\n", encoding="utf-8")

    # Then: callers can load it without knowing the storage layout
    assert temp_storage.load_harmonized_path(file_id) == harmonized_path


async def test_stage5_download_preserves_harmonized_headers(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Download keeps headers produced by harmonization, including Stage 2 column renames."""

    # Given: harmonization wrote a renamed output column
    rows = [["diagnosis"], ["alpha"]]
    file_id = await upload_content(app_client, create_csv_content(rows), "renamed.csv")
    meta = temp_storage.load(file_id)
    assert meta is not None
    harmonized_path = temp_storage.harmonized_path_for(file_id, meta.saved_path)
    with harmonized_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerows([["Primary Diagnosis"], ["Lung Cancer"]])

    # When: the download endpoint is invoked
    response = await app_client.post("/stage-5/download", json={"file_id": file_id})

    # Then: the downloaded CSV keeps the harmonized header and value
    assert response.status_code == 200
    output_rows = _read_downloaded_csv(response.content)
    assert list(output_rows[0]) == ["Primary Diagnosis"]
    assert output_rows[0]["Primary Diagnosis"] == "Lung Cancer"


async def test_stage5_download_succeeds_without_manifest(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Download succeeds even when a manifest is missing."""

    # Given: an uploaded file with a harmonized CSV but no manifest
    rows = [["col_a"], ["alpha"], ["beta"]]
    file_id = await upload_content(app_client, create_csv_content(rows), "no-manifest.csv")
    meta = temp_storage.load(file_id)
    assert meta is not None
    create_harmonized_csv(temp_storage, file_id, meta.saved_path, {})

    # When: the download endpoint is invoked
    response = await app_client.post("/stage-5/download", json={"file_id": file_id})

    # Then: the response contains only the CSV file
    assert response.status_code == 200
    with zipfile.ZipFile(BytesIO(response.content), "r") as zf:
        names = zf.namelist()
    assert any(name.endswith(".csv") for name in names)
    assert not any(name.endswith(".parquet") for name in names)


async def test_stage5_download_includes_manifest_json_when_available(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Download bundles the harmonization manifest as inspectable JSON."""

    # Given: an uploaded file with harmonized output and a stored manifest
    rows = [["col_a"], ["alpha"]]
    file_id = await upload_content(app_client, create_csv_content(rows), "with-manifest.csv")
    meta = temp_storage.load(file_id)
    assert meta is not None
    create_harmonized_csv(temp_storage, file_id, meta.saved_path, {0: {"col_a": "beta"}})
    create_manifest_for_file(temp_storage, file_id, meta.saved_path, {0: {"col_a": "beta"}})

    # When: the download endpoint is invoked
    response = await app_client.post("/stage-5/download", json={"file_id": file_id})

    # Then: the ZIP includes a JSON copy of the manifest rows
    assert response.status_code == 200
    with zipfile.ZipFile(BytesIO(response.content), "r") as zf:
        manifest_name = next(name for name in zf.namelist() if name.endswith("_manifest.json"))
        manifest_rows = json.loads(zf.read(manifest_name).decode("utf-8"))
    assert manifest_rows[0]["column_name"] == "col_a"
    assert manifest_rows[0]["to_harmonize"] == "alpha"
    assert manifest_rows[0]["top_harmonization"] == "beta"


async def test_stage5_download_includes_cde_mapping_artifact(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Download bundles the saved column-to-CDE mapping plan when available."""

    # Given: an uploaded file with a saved CDE mapping document
    rows = [["col_a"], ["alpha"]]
    file_id = await upload_content(app_client, create_csv_content(rows), "with-mapping.csv")
    meta = temp_storage.load(file_id)
    assert meta is not None
    create_harmonized_csv(temp_storage, file_id, meta.saved_path, {})
    dependencies.get_workflow_storage().write_json(
        dependencies.get_user_context(),
        file_id,
        WorkflowFile.CDE_MAPPING,
        {
            "file_id": file_id,
            "generated_at": "2026-05-13T00:00:00+00:00",
            "target_schema": TEST_TARGET_SCHEMA,
            "target_version": "1",
            "mappings": [{"column_key": "col_0000", "source_column_name": "col_a"}],
        },
    )

    # When: the download endpoint is invoked
    response = await app_client.post("/stage-5/download", json={"file_id": file_id})

    # Then: the ZIP includes the mapping artifact alongside the data file
    assert response.status_code == 200
    with zipfile.ZipFile(BytesIO(response.content), "r") as zf:
        mapping_name = next(name for name in zf.namelist() if name.endswith("_cde_mapping.json"))
        mapping_document = json.loads(zf.read(mapping_name).decode("utf-8"))
    assert mapping_document["file_id"] == file_id
    assert mapping_document["mappings"][0]["column_key"] == "col_0000"


async def test_stage5_download_tsv_input_exports_tsv(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """A TSV upload downloads a TSV data file, with comma text preserved."""

    # Given: a TSV input and a TSV-shaped harmonized intermediate
    content = b"col_a\tcol_b\nalpha, beta\tgamma\n"
    file_id = await upload_content(app_client, content, "download.tsv", TEST_TSV_CONTENT_TYPE)
    meta = temp_storage.load(file_id)
    assert meta is not None
    harmonized_path = temp_storage.harmonized_path_for(file_id, meta.saved_path)
    with harmonized_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle, delimiter="\t")
        writer.writerows([["col_a", "col_b"], ["delta, epsilon", "gamma"]])

    # When: the download endpoint is invoked
    response = await app_client.post("/stage-5/download", json={"file_id": file_id})

    # Then: the zip contains TSV output and values are tab-delimited
    assert response.status_code == 200
    with zipfile.ZipFile(BytesIO(response.content), "r") as zf:
        names = zf.namelist()
    assert any(name.endswith(".tsv") for name in names)
    assert not any(name.endswith(".csv") for name in names)
    output_rows = _read_downloaded_tabular(response.content, ".tsv", "\t")
    assert output_rows == [["col_a", "col_b"], ["delta, epsilon", "gamma"]]


async def test_stage5_download_xlsx_input_exports_xlsx_selected_sheet(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """An XLSX upload downloads XLSX output while updating only the selected sheet."""

    # Given: an XLSX input where Stage 1 selected the second sheet
    content = create_xlsx_content({
        "Keep": [["status"], ["unchanged"]],
        "Patients": [["col_a", "col_b"], ["alpha", "gamma"]],
    })
    file_id = await upload_content(app_client, content, "download.xlsx", TEST_XLSX_CONTENT_TYPE)
    analyze_response = await app_client.post(
        "/stage-1/analyze",
        json={
            "file_id": file_id,
            "target_schema": TEST_TARGET_SCHEMA,
            "target_external_version_number": "11.0.4",
            "sheet_name": "Patients",
        },
    )
    assert analyze_response.status_code == 200
    meta = temp_storage.load(file_id)
    assert meta is not None
    harmonized_path = temp_storage.harmonized_path_for(file_id, meta.saved_path)
    harmonized_dataset = dataset_from_rows(
        headers=["col_a", "col_b"],
        rows=[["delta", "gamma"]],
        source_format=TabularFormat.XLSX,
        sheet_name="Patients",
    )
    write_tabular(harmonized_path, harmonized_dataset, template_path=meta.saved_path)

    # When: the download endpoint is invoked
    response = await app_client.post("/stage-5/download", json={"file_id": file_id})

    # Then: the zip contains XLSX output, and non-selected sheets are preserved
    assert response.status_code == 200
    with zipfile.ZipFile(BytesIO(response.content), "r") as zf:
        names = zf.namelist()
    assert any(name.endswith(".xlsx") for name in names)
    assert not any(name.endswith(".csv") for name in names)
    assert _read_downloaded_xlsx(response.content, "Keep") == [["status"], ["unchanged"]]
    assert _read_downloaded_xlsx(response.content, "Patients") == [["col_a", "col_b"], ["delta", "gamma"]]


async def test_stage5_download_ignores_invalid_row_keys(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Overrides for out-of-range row keys do not alter the output."""

    # Given: an uploaded file with harmonized output and invalid overrides
    rows = [["col_a"], ["alpha"], ["beta"]]
    file_id = await upload_content(app_client, create_csv_content(rows), "invalid-rows.csv")
    meta = temp_storage.load(file_id)
    assert meta is not None
    create_harmonized_csv(temp_storage, file_id, meta.saved_path, {})
    create_manifest_for_file(temp_storage, file_id, meta.saved_path, {})
    await app_client.post(
        "/stage-4/overrides",
        json={
            "file_id": file_id,
            "overrides": {
                "99": {"col_0000": {"ai_value": "alpha", "human_value": "gamma", "original_value": "alpha"}},
            },
            "review_state": review_state_payload(),
        },
    )

    # When: the download endpoint is invoked
    response = await app_client.post("/stage-5/download", json={"file_id": file_id})

    # Then: output rows remain unchanged
    assert response.status_code == 200
    output_rows = _read_downloaded_csv(response.content)
    assert output_rows[0]["col_a"] == "alpha"
    assert output_rows[1]["col_a"] == "beta"

async def test_stage5_summary_zero_changes_when_terms_equal(
    app_client: AsyncClient,
    temp_storage: UploadStorage,
) -> None:
    """Summary counts zero AI changes when harmonized values equal originals."""

    # Given: an uploaded file with no changes in the manifest
    rows = [["col_a"], ["alpha"], ["beta"]]
    file_id = await upload_content(app_client, create_csv_content(rows), "summary.csv")
    meta = temp_storage.load(file_id)
    assert meta is not None
    create_harmonized_csv(temp_storage, file_id, meta.saved_path, {})
    create_manifest_for_file(temp_storage, file_id, meta.saved_path, {})

    # When: summary is requested
    response = await app_client.post("/stage-5/summary", json={"file_id": file_id})

    # Then: AI changes are zero
    assert response.status_code == 200
    summary = response.json()
    total_ai_changes = sum(col["ai_changes"] for col in summary["column_summaries"])
    assert total_ai_changes == 0
